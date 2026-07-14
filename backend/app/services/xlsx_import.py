"""Importação de cobertura de calendário a partir da planilha macro do instituto.

A planilha macro tem, por mês, uma aba (ex.: ``Jul-26``) com:
  - uma linha de cabeçalho com os dias do mês (1..31) espalhados em colunas;
  - blocos de turno na coluna A: equipes **A–F** (plantões de 12h), **RM**
    (Reserva Manhã), **RT** (Reserva Tarde), **PIM** (Pátio Manhã) e **PIT**
    (Pátio Tarde). Cada bloco tem uma linha "seção" cujas células indicam qual
    seção cobre aquele turno em cada dia.

Regras de conversão para a cobertura do calendário do DPI (confirmadas com o
gestor), considerando apenas as seções que ele selecionar:

  - ``Plantão 12h[d]``   = nº de equipes A–F cuja seção no dia *d* está selecionada
                           (duas seções no mesmo dia → 2 vagas; soma, não binário).
  - ``Reserva Manhã[d]`` = 1 se a seção de RM estiver selecionada, senão 0.
  - ``Reserva Tarde[d]`` = 1 se a seção de RT estiver selecionada, senão 0.
  - ``Reserva 12h[d]``   = 1 se, em **fim de semana**, RM e RT forem a **mesma**
                           seção (selecionada) — elas se fundem num plantão de
                           reserva de 12h. Nesse caso Manhã/Tarde ficam 0.
  - ``Pátio Manhã[d]``   = 1 se a seção de PIM estiver selecionada, senão 0.
  - ``Pátio Tarde[d]``   = 1 se a seção de PIT estiver selecionada, senão 0.

O fluxo é dividido em duas partes para que a regra viva só no backend:
  1. ``parse_workbook`` lê o arquivo e devolve uma ``grid`` já digerida
     (por dia: seções de plantão + RM/RT/PIM/PIT + se é fim de semana) e a lista
     de seções encontradas — o frontend usa isso para montar os checkboxes.
  2. ``coverage_from_grid`` aplica as regras à ``grid`` dada a seleção do gestor
     e devolve a cobertura por dia/tipo. É uma função pura (sem openpyxl/banco),
     o que a torna trivial de testar com gabarito.
"""
from __future__ import annotations

import calendar as _cal
import datetime
import io
import unicodedata

import openpyxl

from app.core import schedule_types as _st

# Nomes dos tipos de escala (fonte única em core.schedule_types)
T_PLANTAO = _st.PLANTAO_12H
T_RES_M = _st.RESERVA_MANHA
T_RES_T = _st.RESERVA_TARDE
T_RES_12 = _st.RESERVA_12H
T_PAT_M = _st.PATIO_MANHA
T_PAT_T = _st.PATIO_TARDE
ALL_TYPES = _st.ALL_TYPE_NAMES

# Rótulos de bloco na coluna A
PLANTAO_LABELS = ["A", "B", "C", "D", "E", "F"]
RESERVE_LABELS = {"RM", "RT", "PIM", "PIT"}
BLOCK_LABELS = set(PLANTAO_LABELS) | RESERVE_LABELS

# Seções do DPI pré-marcadas por padrão no frontend
DEFAULT_DPI_SECTIONS = ["SPI", "SPBA", "SPD", "SOL", "SPCEF", "SPOIC", "SIV"]

# Valores de célula que não representam uma seção real
NON_SECTION_TOKENS = {"DISPENSADA", "DISPENSADO", "FOLGA", "-", "X", "XX"}

MONTH_ABBR_PT = {
    1: "Jan", 2: "Fev", 3: "Mar", 4: "Abr", 5: "Mai", 6: "Jun",
    7: "Jul", 8: "Ago", 9: "Set", 10: "Out", 11: "Nov", 12: "Dez",
}


def sheet_name_for(year: int, month: int) -> str:
    """(2026, 7) -> 'Jul-26'."""
    return f"{MONTH_ABBR_PT[month]}-{year % 100:02d}"


def _strip_accents(s: str) -> str:
    return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()


def _norm_label(v) -> str:
    """Normaliza rótulo da coluna A: 'A\\n(12h)' -> 'A'."""
    if v is None:
        return ""
    s = str(v).strip().split("\n")[0].split("(")[0].strip()
    return s.upper()


def _norm_section(v) -> str:
    if v is None:
        return ""
    return str(v).strip().upper()


class SheetLayout:
    """Detecta a geometria da aba (linha dos dias, colunas por dia e linhas de bloco),
    tolerando pequenas variações de layout entre os meses."""

    def __init__(self, ws):
        self.ws = ws
        self.day_header_row, self.day_cols = self._find_day_header(ws)
        self.block_rows = self._find_block_rows(ws, self.day_header_row)

    @staticmethod
    def _find_day_header(ws):
        best = None
        max_col = min(ws.max_column, 80)
        for r in range(1, 12):
            cols = {}
            for c in range(1, max_col + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, int) and 1 <= v <= 31:
                    cols[v] = c
            # exige a sequência inicial 1..5 para caracterizar a linha de dias
            if all(d in cols for d in range(1, 6)):
                if best is None or len(cols) > len(best[1]):
                    best = (r, cols)
        if best is None:
            raise ValueError("Não foi possível localizar a linha com os dias do mês na planilha.")
        return best

    @staticmethod
    def _find_block_rows(ws, header_row):
        rows: dict[str, int] = {}
        for r in range(header_row + 1, min(header_row + 30, ws.max_row + 1)):
            label = _norm_label(ws.cell(row=r, column=1).value)
            colb = _strip_accents(_norm_section(ws.cell(row=r, column=2).value))
            # A linha "seção" traz o rótulo na col A e a palavra 'seção' na col B
            if label in BLOCK_LABELS and colb.startswith("SECAO"):
                rows.setdefault(label, r)
        if not rows:
            raise ValueError("Não foi possível localizar os blocos de turno (A–F, RM, RT, PIM, PIT).")
        return rows


def build_grid(ws, year: int, month: int) -> list[dict]:
    """Digere a aba num formato leve por dia — sem aplicar seleção ainda."""
    layout = SheetLayout(ws)
    _, ndays = _cal.monthrange(year, month)
    days = [d for d in range(1, ndays + 1) if d in layout.day_cols]
    br = layout.block_rows

    def sec(label: str, d: int) -> str:
        row = br.get(label)
        if not row:
            return ""
        return _norm_section(ws.cell(row=row, column=layout.day_cols[d]).value)

    grid: list[dict] = []
    for d in days:
        plantao = [sec(lbl, d) for lbl in PLANTAO_LABELS]
        plantao = [s for s in plantao if s]  # remove equipes vazias
        grid.append({
            "day": d,
            "weekend": datetime.date(year, month, d).weekday() >= 5,
            "plantao": plantao,
            "rm": sec("RM", d),
            "rt": sec("RT", d),
            "pim": sec("PIM", d),
            "pit": sec("PIT", d),
        })
    return grid


def sections_in_grid(grid: list[dict]) -> list[str]:
    """Todas as seções reais encontradas na grade (exclui DISPENSADA etc.)."""
    found: set[str] = set()
    for cell in grid:
        for s in list(cell["plantao"]) + [cell["rm"], cell["rt"], cell["pim"], cell["pit"]]:
            s = _norm_section(s)
            if s and s not in NON_SECTION_TOKENS:
                found.add(s)
    return sorted(found)


def coverage_from_grid(grid: list[dict], selected_sections) -> dict[int, dict[str, int]]:
    """Aplica as regras à grade dada a seleção de seções. Função pura e testável.

    Retorna ``{dia: {nome_tipo: quantidade}}`` com os 6 tipos sempre presentes.
    """
    sel = {_norm_section(s) for s in selected_sections if _norm_section(s)}
    out: dict[int, dict[str, int]] = {}
    for cell in grid:
        d = cell["day"]
        cov = {t: 0 for t in ALL_TYPES}

        # Plantão 12h: soma das equipes A–F cuja seção está selecionada
        cov[T_PLANTAO] = sum(1 for s in cell["plantao"] if _norm_section(s) in sel)

        rm, rt = _norm_section(cell["rm"]), _norm_section(cell["rt"])
        if cell["weekend"] and rm and rm == rt:
            # Fim de semana com RM e RT da mesma seção → Reserva 12h
            cov[T_RES_12] = 1 if rm in sel else 0
        else:
            cov[T_RES_M] = 1 if rm in sel else 0
            cov[T_RES_T] = 1 if rt in sel else 0

        cov[T_PAT_M] = 1 if _norm_section(cell["pim"]) in sel else 0
        cov[T_PAT_T] = 1 if _norm_section(cell["pit"]) in sel else 0

        out[d] = cov
    return out


def parse_workbook(file_bytes: bytes, year: int, month: int, sheet_name: str | None = None) -> dict:
    """Lê o arquivo xlsx e devolve a grade digerida + seções encontradas.

    Levanta ``ValueError`` se a aba do mês não existir ou o layout não for reconhecido.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    name = sheet_name or sheet_name_for(year, month)
    if name not in wb.sheetnames:
        disponiveis = ", ".join(wb.sheetnames[:16])
        raise ValueError(f"Aba '{name}' não encontrada na planilha. Abas disponíveis: {disponiveis}")
    ws = wb[name]
    grid = build_grid(ws, year, month)
    sections = sections_in_grid(grid)
    default_selected = [s for s in DEFAULT_DPI_SECTIONS if s in sections]
    return {
        "sheet_name": name,
        "available_sheets": wb.sheetnames,
        "grid": grid,
        "sections_found": sections,
        "default_selected": default_selected,
        "types": ALL_TYPES,
    }
